import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from typing import Dict, Any, Optional, Union, List, Set, Tuple
from datetime import datetime
from enum import Enum
import os
import threading
import time
from src.utils.logger import get_logger, log_metric, correlation_decorator
from src.utils.config_loader import ConfigLoader


class CurrencySymbol(Enum):
    USD = "$"
    EUR = "€"
    GBP = "£"
    JPY = "¥JPY"
    INR = "₹"
    CNY = "¥CNY"
    AUD = "A$"
    CAD = "C$"


class ExcelExporter:

    def __init__(self, export_dir: str = "data/exports", currency: str = "USD"):
        self.logger = get_logger("ExcelExporter")
        self.export_dir = Path(export_dir)
        self.lock = threading.RLock()
        self.template_path = None
        self.append_mode = False

        try:
            self.currency = CurrencySymbol[currency.upper()].value
        except (KeyError, AttributeError):
            self.currency = "$"
            self.logger.warning(f"Unknown currency '{currency}', defaulting to USD")

        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")

        self.dark_blue_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
        self.light_fill = PatternFill(start_color="E7F0F9", end_color="E7F0F9", fill_type="solid")
        self.white_fill = PatternFill(fill_type=None)

        self.data_font = Font(size=10, name="Calibri", color="000000")
        self.header_font_alt = Font(size=9, name="Calibri", color="666666")

        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
        self.right_align = Alignment(horizontal="right", vertical="center", wrap_text=False)

        self.thin_border = Side(border_style="thin", color="404040")
        self.thin_border_light = Side(border_style="thin", color="D3D3D3")

        try:
            self._ensure_export_directory()
            self.logger.info(
                f"ExcelExporter initialized (export_dir={str(self.export_dir)}, "
                f"currency={self.currency}, append_mode={self.append_mode})"
            )
            log_metric(
                "excel_exporter_init",
                1,
                {
                    "export_dir": str(self.export_dir),
                    "currency": currency,
                    "append_mode": self.append_mode,
                },
            )
        except Exception as e:
            self.logger.error(f"Failed to initialize ExcelExporter: {str(e)}")
            log_metric("excel_exporter_init", 0, {"error": str(e)})
            raise

    def set_template(self, template_path: str) -> bool:
        try:
            path = Path(template_path)
            if not path.exists():
                self.logger.warning(f"Template file not found: {template_path}")
                return False
            self.template_path = template_path
            self.logger.info(f"Template set: {template_path}")
            return True
        except Exception as e:
            self.logger.error(f"Error setting template: {str(e)}")
            return False

    def enable_append_mode(self, enable: bool = True) -> None:
        self.append_mode = enable
        self.logger.info(f"Append mode: {enable}")

    def _ensure_export_directory(self) -> None:
        try:
            self.export_dir.mkdir(parents=True, exist_ok=True)
        except (OSError, PermissionError) as e:
            self.logger.error(f"Failed to create export directory: {str(e)}")
            raise

    def _apply_number_formatting(self, ws: Any, df: pd.DataFrame) -> None:
        try:
            if ws.max_row < 2:
                return

            for col_idx, col in enumerate(df.columns, 1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter

                if any(keyword in str(col).lower() for keyword in ["price", "cost", "amount", "value"]):
                    for row_idx in range(2, ws.max_row + 1):
                        try:
                            cell = ws[f"{col_letter}{row_idx}"]
                            cell.number_format = f"{self.currency} #,##0.00"
                        except Exception:
                            pass

                elif any(keyword in str(col).lower() for keyword in ["volume", "count"]):
                    for row_idx in range(2, ws.max_row + 1):
                        try:
                            cell = ws[f"{col_letter}{row_idx}"]
                            cell.number_format = "#,##0"
                        except Exception:
                            pass

                elif any(keyword in str(col).lower() for keyword in ["percent", "change", "ratio"]):
                    for row_idx in range(2, ws.max_row + 1):
                        try:
                            cell = ws[f"{col_letter}{row_idx}"]
                            cell.number_format = "0.00%"
                        except Exception:
                            pass

        except Exception as e:
            self.logger.debug(f"Error applying number formatting: {str(e)}")

    def _apply_conditional_formatting(self, ws: Any) -> None:
        try:
            if ws.max_row < 2:
                return

            from openpyxl.formatting.rule import CellIsRule

            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for col_idx in range(1, ws.max_column + 1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter

                try:
                    col_header = str(ws[f"{col_letter}1"].value or "").lower()
                    if any(k in col_header for k in ["change", "return", "gain"]):
                        green_rule = CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill)
                        red_rule = CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)

                        range_str = f"{col_letter}2:{col_letter}{ws.max_row}"
                        ws.conditional_formatting.add(range_str, green_rule)
                        ws.conditional_formatting.add(range_str, red_rule)
                except Exception:
                    pass

        except Exception as e:
            self.logger.debug(f"Error applying conditional formatting: {str(e)}")

    def _apply_data_validation(self, ws: Any, df: pd.DataFrame) -> None:
        try:
            if ws.max_row < 2:
                return

            from openpyxl.worksheet.datavalidation import DataValidation

            for col_idx, col in enumerate(df.columns, 1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter

                if any(k in str(col).lower() for k in ["type", "asset", "category"]):
                    dv = DataValidation(
                        type="list",
                        formula1='"stock,crypto,bond,commodity,forex"',
                        allow_blank=True
                    )
                    dv.error = "Select from list"
                    dv.errorTitle = "Invalid Entry"
                    ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")

        except Exception as e:
            self.logger.debug(f"Error applying data validation: {str(e)}")

    def _apply_freeze_panes(self, ws: Any) -> None:
        try:
            ws.freeze_panes = "A2"
        except Exception as e:
            self.logger.debug(f"Error applying freeze panes: {str(e)}")

    def _apply_print_settings(self, ws: Any) -> None:
        try:
            ws.page_setup.orientation = "landscape"
            ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.75
            ws.page_margins.bottom = 0.75

            ws.print_options.horizontalCentered = False
            ws.print_title_rows = "1:1"

        except Exception as e:
            self.logger.debug(f"Error applying print settings: {str(e)}")

    def _to_dataframe(self, data: Any) -> pd.DataFrame:
        try:
            if isinstance(data, pd.DataFrame):
                return data.copy()

            if isinstance(data, dict):
                if not data:
                    return pd.DataFrame()

                if isinstance(next(iter(data.values()), None), (dict, list)):
                    df = pd.DataFrame.from_dict(data, orient='index')
                    df = df.reset_index()
                    df.columns = self._clean_column_names(df.columns)
                    return df
                else:
                    df = pd.DataFrame([data])
                    df.columns = self._clean_column_names(df.columns)
                    return df

            if isinstance(data, list):
                if not data:
                    return pd.DataFrame()
                df = pd.DataFrame(data)
                df.columns = self._clean_column_names(df.columns)
                return df

            self.logger.warning(f"Unsupported data format: {type(data)}")
            return pd.DataFrame()

        except Exception as e:
            self.logger.error(f"Error converting data to DataFrame: {str(e)}")
            return pd.DataFrame()
    
    def _clean_column_names(self, columns: List[str]) -> List[str]:
        cleaned = []
        for col in columns:
            try:
                col_str = str(col).strip()
                
                if not col_str:
                    cleaned.append("Column")
                    continue
                
                col_str = col_str.replace('_', ' ').replace('-', ' ')
                words = col_str.split()
                formatted = ' '.join(word.capitalize() for word in words if word)
                cleaned.append(formatted if formatted else "Column")
            except Exception:
                cleaned.append(str(col))
        
        return cleaned

    def _validate_dataframe(self, df: pd.DataFrame) -> bool:
        try:
            if df is None or not isinstance(df, pd.DataFrame):
                return False

            if df.empty:
                self.logger.debug("DataFrame is empty")
                return True

            if len(df.columns) == 0:
                self.logger.warning("DataFrame has no columns")
                return False

            return True

        except (TypeError, AttributeError, ValueError) as e:
            self.logger.error(f"DataFrame validation error: {str(e)}")
            return False

    def _write_dataframe_to_sheet(self, df: pd.DataFrame, ws: Any) -> int:
        try:
            if not self._validate_dataframe(df):
                return 0

            row_count = 0
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    try:
                        if r_idx == 1:
                            cell.fill = self.header_fill
                            cell.font = self.header_font
                            cell.alignment = self.center_align
                            cell.border = Border(
                                left=self.thin_border,
                                right=self.thin_border,
                                top=self.thin_border,
                                bottom=self.thin_border
                            )
                        else:
                            cell.font = self.data_font
                            
                            if (r_idx - 2) % 2 == 0:
                                cell.fill = self.dark_blue_fill
                            else:
                                cell.fill = self.white_fill
                            
                            cell.border = Border(
                                left=self.thin_border_light,
                                right=self.thin_border_light,
                                top=self.thin_border_light,
                                bottom=self.thin_border_light
                            )
                        
                        if value is None:
                            cell.value = ""
                            cell.alignment = self.center_align
                        elif isinstance(value, bool):
                            cell.value = str(value)
                            cell.alignment = self.center_align
                        elif isinstance(value, (int, float)):
                            cell.value = value
                            cell.alignment = self.right_align
                        elif isinstance(value, datetime):
                            cell.value = value.isoformat()
                            cell.alignment = self.center_align
                        else:
                            cell_value = str(value)[:32767]
                            cell.value = cell_value if cell_value else ""
                            cell.alignment = self.center_align
                    except (ValueError, TypeError, AttributeError) as e:
                        self.logger.debug(f"Error writing cell ({r_idx}, {c_idx}): {str(e)}")
                        cell.value = ""
                        cell.alignment = self.center_align

                row_count = r_idx

            return row_count

        except (ValueError, TypeError, AttributeError) as e:
            self.logger.error(f"Error writing DataFrame to sheet: {str(e)}")
            return 0

    def _apply_header_style(self, ws: Any) -> None:
        try:
            if ws.max_row < 1:
                return

            for cell in ws[1]:
                try:
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.alignment = self.center_align
                    cell.border = Border(
                        left=self.thin_border,
                        right=self.thin_border,
                        top=self.thin_border,
                        bottom=self.thin_border
                    )
                except (ValueError, TypeError, AttributeError) as e:
                    self.logger.debug(f"Error styling header cell: {str(e)}")

        except (ValueError, TypeError, AttributeError) as e:
            self.logger.error(f"Error applying header style: {str(e)}")

    def _apply_row_colors(self, ws: Any) -> None:
        try:
            if ws.max_row < 2:
                return

            for row_idx in range(2, ws.max_row + 1):
                row = ws[row_idx]
                if (row_idx - 2) % 2 == 0:
                    fill = self.dark_blue_fill
                else:
                    fill = self.white_fill
                
                for cell in row:
                    try:
                        cell.fill = fill
                        cell.font = self.data_font
                        cell.alignment = self.center_align
                        cell.border = Border(
                            left=self.thin_border_light,
                            right=self.thin_border_light,
                            top=self.thin_border_light,
                            bottom=self.thin_border_light
                        )
                    except (ValueError, TypeError, AttributeError) as e:
                        self.logger.debug(f"Error applying row color: {str(e)}")

        except (ValueError, TypeError, AttributeError) as e:
            self.logger.error(f"Error applying row colors: {str(e)}")

    def _get_unique_sheet_name(self, sheet_name: str, existing_names: Set[str]) -> str:
        truncated = sheet_name[:31]
        
        if truncated not in existing_names:
            return truncated
        
        base_name = sheet_name[:25]
        counter = 2
        
        while counter <= 1000:
            candidate = f"{base_name} ({counter})"
            if len(candidate) <= 31 and candidate not in existing_names:
                return candidate
            counter += 1
        
        return sheet_name[:31]

    def _apply_thin_borders(self, ws: Any) -> None:
        try:
            if ws.max_row < 1 or ws.max_column < 1:
                return

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    try:
                        cell.border = Border(
                            left=self.thin_border_light if cell.row > 1 else self.thin_border,
                            right=self.thin_border_light if cell.row > 1 else self.thin_border,
                            top=self.thin_border_light if cell.row > 1 else self.thin_border,
                            bottom=self.thin_border_light if cell.row > 1 else self.thin_border
                        )
                    except (ValueError, TypeError, AttributeError) as e:
                        self.logger.debug(f"Error applying border: {str(e)}")

        except (ValueError, TypeError, AttributeError) as e:
            self.logger.error(f"Error applying borders: {str(e)}")

    def _auto_column_width(self, ws: Any) -> None:
        try:
            if ws.max_column < 1:
                return

            for col_idx in range(1, ws.max_column + 1):
                try:
                    column = ws.cell(row=1, column=col_idx).column_letter
                    max_len = 12
                    col_cells = ws[column]

                    for cell in col_cells:
                        try:
                            if cell.value:
                                length = len(str(cell.value))
                                max_len = max(max_len, min(length, 50))
                        except (ValueError, TypeError, AttributeError) as e:
                            self.logger.debug(f"Error calculating column width: {str(e)}")

                    adjusted_width = min(max_len + 3, 60)
                    ws.column_dimensions[column].width = adjusted_width

                except (ValueError, TypeError, AttributeError, KeyError) as e:
                    self.logger.debug(f"Error auto-sizing column {col_idx}: {str(e)}")

        except (ValueError, TypeError, AttributeError) as e:
            self.logger.error(f"Error applying auto column width: {str(e)}")

    def _apply_enterprise_styling(self, ws: Any, df: pd.DataFrame) -> None:
        try:
            self._apply_header_style(ws)
            self._apply_row_colors(ws)
            self._apply_thin_borders(ws)
            self._auto_column_width(ws)
            self._set_row_height(ws)
            self._apply_number_formatting(ws, df)
            self._apply_conditional_formatting(ws)
            self._apply_data_validation(ws, df)
            self._apply_freeze_panes(ws)
            self._apply_print_settings(ws)
        except (ValueError, TypeError, AttributeError) as e:
            self.logger.error(f"Error applying enterprise styling: {str(e)}")

    def _set_row_height(self, ws: Any) -> None:
        try:
            for row_idx in range(1, ws.max_row + 1):
                try:
                    if row_idx == 1:
                        ws.row_dimensions[row_idx].height = 24
                    else:
                        ws.row_dimensions[row_idx].height = 20
                except (ValueError, TypeError, AttributeError, KeyError):
                    pass
        except (ValueError, TypeError, AttributeError) as e:
            self.logger.debug(f"Error setting row height: {str(e)}")

    def _save(self, wb: Workbook, prefix: str, max_retries: int = 3, retry_delay: float = 0.5) -> Optional[str]:
        try:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            base_filename = f"{prefix}_{timestamp}.xlsx"
            filepath = self.export_dir / base_filename
            
            with self.lock:
                for attempt in range(max_retries):
                    try:
                        wb.save(str(filepath))
                        self.logger.info(f"Excel file saved: {str(filepath)}")
                        log_metric("excel_save_success", 1, {"filename": base_filename, "path": str(filepath)})
                        return str(filepath)
                    
                    except (IOError, OSError, PermissionError) as e:
                        if attempt < max_retries - 1:
                            self.logger.warning(f"Save attempt {attempt + 1} failed, retrying in {retry_delay}s: {str(e)}")
                            time.sleep(retry_delay)
                        else:
                            raise
                
                fallback_filepath = self._generate_unique_filepath(prefix, timestamp)
                wb.save(str(fallback_filepath))
                self.logger.info(f"Excel file saved with fallback path: {str(fallback_filepath)}")
                log_metric("excel_save_success_fallback", 1, {"filename": fallback_filepath.name, "path": str(fallback_filepath)})
                return str(fallback_filepath)

        except IOError as e:
            self.logger.error(f"IO error saving Excel file: {str(e)}")
            log_metric("excel_save_error", 0, {"error": f"IOError: {str(e)}"})
            return None
        except Exception as e:
            self.logger.error(f"Error saving Excel file: {str(e)}")
            log_metric("excel_save_error", 0, {"error": str(e)})
            return None
    
    def _generate_unique_filepath(self, prefix: str, timestamp: str) -> Path:
        counter = 1
        while True:
            unique_filename = f"{prefix}_{timestamp}_{counter}.xlsx"
            filepath = self.export_dir / unique_filename
            if not filepath.exists():
                return filepath
            counter += 1
            if counter > 1000:
                raise RuntimeError("Unable to generate unique filename after 1000 attempts")

    @correlation_decorator()
    def export_to_excel(self, data: Any, filename_prefix: str = "market_data") -> Optional[str]:
        try:
            if data is None:
                self.logger.warning("No data provided for export")
                log_metric("excel_export_no_data", 0, {})
                return None

            df = self._to_dataframe(data)

            if not self._validate_dataframe(df):
                self.logger.error("Invalid DataFrame for export")
                log_metric("excel_export_invalid_df", 0, {})
                return None

            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Data"

                row_count = self._write_dataframe_to_sheet(df, ws)

                if row_count == 0:
                    self.logger.warning("No rows written to sheet")
                    log_metric("excel_export_no_rows", 0, {})
                    return None

                self._apply_enterprise_styling(ws, df)

                filepath = self._save(wb, filename_prefix)

                if filepath:
                    log_metric(
                        "excel_export_success",
                        1,
                        {
                            "filename_prefix": filename_prefix,
                            "rows": row_count,
                            "columns": df.shape[1],
                        },
                    )

                return filepath

            except Exception as e:
                self.logger.error(f"Error during export_to_excel: {str(e)}")
                log_metric("excel_export_error", 0, {"error": str(e)})
                return None

        except Exception as e:
            self.logger.error(f"Unexpected error in export_to_excel: {str(e)}")
            log_metric("excel_export_unexpected_error", 0, {"error": str(e)})
            return None

    @correlation_decorator()
    def export_multi_sheet(self, data_dict: Dict[str, Any], filename_prefix: str = "market_data") -> Optional[str]:
        try:
            if not data_dict or not isinstance(data_dict, dict):
                self.logger.warning("Invalid data_dict for multi-sheet export")
                log_metric("excel_multi_sheet_invalid", 0, {})
                return None

            try:
                wb = Workbook()
                wb.remove(wb.active)

                sheet_count = 0
                total_rows = 0
                used_sheet_names: Set[str] = set()

                for sheet_name, data in data_dict.items():
                    try:
                        if not sheet_name or not isinstance(sheet_name, str):
                            self.logger.warning(f"Invalid sheet name: {sheet_name}")
                            continue

                        df = self._to_dataframe(data)

                        if not self._validate_dataframe(df):
                            self.logger.warning(f"Invalid DataFrame for sheet: {sheet_name}")
                            continue

                        unique_sheet_name = self._get_unique_sheet_name(sheet_name, used_sheet_names)
                        used_sheet_names.add(unique_sheet_name)
                        
                        ws = wb.create_sheet(title=unique_sheet_name)

                        row_count = self._write_dataframe_to_sheet(df, ws)

                        if row_count > 0:
                            self._apply_enterprise_styling(ws, df)
                            sheet_count += 1
                            total_rows += row_count
                        else:
                            self.logger.warning(f"No rows written to sheet: {sheet_name}")
                            wb.remove(ws)

                    except Exception as e:
                        self.logger.error(f"Error processing sheet '{sheet_name}': {str(e)}")
                        continue

                if sheet_count == 0:
                    self.logger.error("No sheets were successfully created")
                    log_metric("excel_multi_sheet_no_sheets", 0, {})
                    return None

                filepath = self._save(wb, filename_prefix)

                if filepath:
                    log_metric("excel_multi_sheet_success", 1, {
                        "sheets": sheet_count,
                        "total_rows": total_rows,
                        "filename_prefix": filename_prefix
                    })

                return filepath

            except Exception as e:
                self.logger.error(f"Error during export_multi_sheet: {str(e)}")
                log_metric("excel_multi_sheet_error", 0, {"error": str(e)})
                return None

        except Exception as e:
            self.logger.error(f"Unexpected error in export_multi_sheet: {str(e)}")
            log_metric("excel_multi_sheet_unexpected_error", 0, {"error": str(e)})
            return None

    def health_check(self) -> Dict[str, Any]:
        try:
            health = {
                "status": "healthy",
                "export_dir": str(self.export_dir),
                "export_dir_exists": self.export_dir.exists(),
                "export_dir_writable": os.access(str(self.export_dir), os.W_OK),
                "timestamp": datetime.utcnow().isoformat()
            }
            self.logger.debug("Health check: OK")
            log_metric("excel_exporter_health_check", 1, health)
            return health

        except (OSError, ValueError, TypeError, AttributeError) as e:
            self.logger.error(f"Health check failed: {str(e)}")
            log_metric("excel_exporter_health_check", 0, {"error": str(e)})
            return {"status": "unhealthy", "error": str(e)}

    def __del__(self):
        try:
            pass
        except Exception as e:
            self.logger.error(f"Error in destructor: {str(e)}")


class MLExcelHelper:

    @staticmethod
    def prepare_prediction_sheet(predictions: Any) -> pd.DataFrame:
        try:
            if isinstance(predictions, pd.DataFrame):
                return predictions.copy()

            if isinstance(predictions, dict):
                return pd.DataFrame.from_dict(predictions, orient='index').reset_index(drop=True)

            if isinstance(predictions, list):
                return pd.DataFrame(predictions)

            return pd.DataFrame()

        except (ValueError, TypeError, KeyError, AttributeError) as e:
            logger = get_logger("MLExcelHelper")
            logger.error(f"Error preparing prediction sheet: {str(e)}")
            return pd.DataFrame()

    @staticmethod
    def prepare_feature_sheet(feature_importance: Any) -> pd.DataFrame:
        try:
            if isinstance(feature_importance, pd.DataFrame):
                return feature_importance.copy()

            if isinstance(feature_importance, dict):
                return pd.DataFrame(list(feature_importance.items()), columns=["Feature", "Importance"])

            if isinstance(feature_importance, list):
                return pd.DataFrame(feature_importance)

            return pd.DataFrame()

        except (ValueError, TypeError, KeyError, AttributeError) as e:
            logger = get_logger("MLExcelHelper")
            logger.error(f"Error preparing feature sheet: {str(e)}")
            return pd.DataFrame()

    @staticmethod
    def prepare_metadata_sheet(model_info: Dict[str, Any]) -> pd.DataFrame:
        try:
            if not isinstance(model_info, dict):
                return pd.DataFrame()

            metadata_df = pd.DataFrame([{
                "Key": k,
                "Value": str(v)[:32767]
            } for k, v in model_info.items()])

            return metadata_df

        except (ValueError, TypeError, KeyError, AttributeError) as e:
            logger = get_logger("MLExcelHelper")
            logger.error(f"Error preparing metadata sheet: {str(e)}")
            return pd.DataFrame()

    @staticmethod
    def _format_pair_name(pair_key: str) -> str:
        try:
            if not pair_key or not isinstance(pair_key, str):
                return str(pair_key)
            
            parts = pair_key.split('_')
            formatted_parts = [part.capitalize() for part in parts]
            return ' '.join(formatted_parts)
        except (AttributeError, ValueError):
            return pair_key

    @staticmethod
    def _format_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
        try:
            if df is None or df.empty:
                return df
            
            column_mapping = {}
            for col in df.columns:
                if isinstance(col, str):
                    if '_' in col:
                        column_mapping[col] = MLExcelHelper._format_pair_name(col)
                    elif col and col[0].islower():
                        column_mapping[col] = col[0].upper() + col[1:]
            
            if column_mapping:
                df = df.rename(columns=column_mapping)
            
            return df
        except (ValueError, TypeError, AttributeError, KeyError) as e:
            logger = get_logger("MLExcelHelper")
            logger.warning(f"Error formatting DataFrame columns: {str(e)}")
            return df

    @staticmethod
    def _split_timestamp_value(timestamp_str: str) -> Tuple[str, str]:
        try:
            if not timestamp_str or not isinstance(timestamp_str, str):
                return "", ""
            
            if "T" in timestamp_str:
                date_part, time_part = timestamp_str.split("T", 1)
                time_clean = time_part.replace("Z", "").split(".")[0]
                return date_part, time_clean
            elif " " in timestamp_str:
                parts = timestamp_str.split(" ", 1)
                date_part = parts[0]
                time_clean = parts[1].split(".")[0] if len(parts) > 1 else ""
                return date_part, time_clean
            else:
                return timestamp_str, ""
        except (ValueError, AttributeError, IndexError):
            return timestamp_str, ""

    @staticmethod
    def _process_sheet_timestamps(df: pd.DataFrame) -> pd.DataFrame:
        try:
            if df is None or df.empty:
                return df
            
            timestamp_col = None
            for col in df.columns:
                if isinstance(col, str) and col.lower() == "timestamp":
                    timestamp_col = col
                    break
            
            if timestamp_col is None:
                return df
            
            date_values = []
            time_values = []
            
            for val in df[timestamp_col]:
                date_val, time_val = MLExcelHelper._split_timestamp_value(str(val))
                date_values.append(date_val)
                time_values.append(time_val)
            
            df_copy = df.copy()
            df_copy = df_copy.drop(columns=[timestamp_col])
            df_copy.insert(len(df_copy.columns), "Date", date_values)
            df_copy.insert(len(df_copy.columns), "Time", time_values)
            
            return df_copy
        except (ValueError, TypeError, KeyError, AttributeError) as e:
            logger = get_logger("MLExcelHelper")
            logger.warning(f"Error processing sheet timestamps: {str(e)}")
            return df

    @staticmethod
    def prepare_ml_ready_data(aggregated_data: Dict[str, Any]) -> Dict[str, Any]:
        try:
            if not aggregated_data or not isinstance(aggregated_data, dict):
                return {}

            symbols = aggregated_data.get("symbols", {})
            if not symbols:
                return {}

            data_rows = []
            exporter = ExcelExporter()
            
            for symbol, data in symbols.items():
                try:
                    if not isinstance(data, dict):
                        continue
                        
                    norm = data.get("normalized", {})
                    ml = data.get("ml_features", {})
                    tech = data.get("technical_indicators", {})
                    basic = data.get("basic_features", {})
                    kline = data.get("kline_features", {})
                    
                    asset_type = norm.get("type", "stock").lower()
                    
                    timestamp_raw = norm.get("timestamp", "")
                    date_part, time_clean = MLExcelHelper._split_timestamp_value(str(timestamp_raw)) if timestamp_raw else ("", "")

                    price_val = norm.get("price", 0)
                    high_val = norm.get("high", 0)
                    low_val = norm.get("low", 0)
                    open_val = norm.get("open", 0)

                    row = {
                        "Symbol": symbol,
                        "Type": norm.get("type", "N/A").upper(),
                        "Price": exporter._format_price(price_val, asset_type),
                        "Change": f"{exporter.currency}{round(norm.get('change', 0), 2):.2f}",
                        "Change %": f"{round(norm.get('change_percent', 0), 2):.2f}%",
                        "Volume": f"{int(norm.get('volume', 0)):,}",
                        "High": exporter._format_price(high_val, asset_type),
                        "Low": exporter._format_price(low_val, asset_type),
                        "Open": exporter._format_price(open_val, asset_type),
                        "Momentum": round(ml.get("momentum", 0), 2),
                        "Trend Strength": round(ml.get("trend_strength", 0), 2),
                        "Volatility %": f"{round(ml.get('volatility_normalized', 0), 2):.2f}%",
                        "Volume Norm": round(ml.get("volume_normalized", 0), 2),
                        "Price Position": round(ml.get("price_position_ratio", 0.5), 2),
                        "Data Quality": round(ml.get("data_quality_score", 0), 2),
                        "RSI 14": round(tech.get("rsi_14", 0), 2),
                        "MACD": round(tech.get("macd", 0), 4),
                        "ATR": round(tech.get("atr", 0), 2),
                        "EMA 12": exporter._format_price(tech.get("ema_12", 0), asset_type),
                        "EMA 26": exporter._format_price(tech.get("ema_26", 0), asset_type),
                        "Price Range": exporter._format_price(basic.get("price_range", 0), asset_type),
                        "HL Ratio": round(basic.get("hl_ratio", 1.0), 2),
                        "Date": date_part,
                        "Time": time_clean
                    }

                    if kline and isinstance(kline, dict):
                        row.update({
                            "Rolling Avg 24h": exporter._format_price(kline.get("rolling_mean_24h", 0), asset_type),
                            "Rolling Std 24h": exporter._format_price(kline.get("rolling_std_24h", 0), asset_type),
                            "Volume Change 24h": f"{round(kline.get('rolling_volume_change', 0), 2):.2f}%",
                            "Kline Count": int(kline.get("kline_count", 0))
                        })

                    data_rows.append(row)

                except (TypeError, ValueError, KeyError, AttributeError) as e:
                    logger = get_logger("MLExcelHelper")
                    logger.warning(f"Error preparing ML data for {symbol}: {str(e)}")
                    continue

            if not data_rows:
                return {}

            market_summary = aggregated_data.get("market_summary", {})
            correlations = aggregated_data.get("correlations", {})

            market_data_df = pd.DataFrame(data_rows)
            
            market_summary_rows = []
            for k, v in market_summary.items():
                if k.lower() == "timestamp":
                    date_val, time_val = MLExcelHelper._split_timestamp_value(str(v))
                    market_summary_rows.append({"Metric": "Date", "Value": date_val})
                    market_summary_rows.append({"Metric": "Time", "Value": time_val})
                else:
                    market_summary_rows.append({
                        "Metric": MLExcelHelper._format_pair_name(k),
                        "Value": v
                    })
            
            market_summary_df = pd.DataFrame(market_summary_rows)
            
            correlations_rows = []
            for k, v in correlations.items():
                if k.lower() == "timestamp":
                    date_val, time_val = MLExcelHelper._split_timestamp_value(str(v))
                    correlations_rows.append({"Pair": "Date", "Correlation": date_val})
                    correlations_rows.append({"Pair": "Time", "Correlation": time_val})
                else:
                    correlations_rows.append({
                        "Pair": MLExcelHelper._format_pair_name(k),
                        "Correlation": v
                    })
            
            correlations_df = pd.DataFrame(correlations_rows)

            market_data_df = MLExcelHelper._format_dataframe_columns(market_data_df)
            market_summary_df = MLExcelHelper._format_dataframe_columns(market_summary_df)
            correlations_df = MLExcelHelper._format_dataframe_columns(correlations_df)

            return {
                "Market Data": market_data_df,
                "Market Summary": market_summary_df,
                "Correlations": correlations_df
            }

        except (ValueError, TypeError, KeyError, AttributeError) as e:
            logger = get_logger("MLExcelHelper")
            logger.error(f"Error preparing ML-ready data: {str(e)}")
            return {}
